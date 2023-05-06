from urllib.request import urlopen, urlretrieve
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, colors
from openpyxl import workbook
from openpyxl.drawing.image import Image
import keys
from twilio.rest import Client

###before running this code it is important to first import Pillow###

wb = xl.Workbook()
ws = wb.active
ws.title = 'Top Cryptocurrencies'
sheet = wb['Top Cryptocurrencies']
ws.showGridlines= False

ws['B1'] = 'Currency'
ws['C1'] = 'Current Value'
ws['D1'] = 'Percent Change'

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 30

ws.row_dimensions[1].height = 60
ws.row_dimensions[2].height = 180
ws.row_dimensions[3].height = 180
ws.row_dimensions[4].height = 180
ws.row_dimensions[5].height = 180
ws.row_dimensions[6].height = 180

cell_range = ws['B2:D6']
for row in cell_range:
    for cell in row:
        cell.alignment = xl.styles.Alignment(horizontal='center',vertical='center')
        cell.font = Font(name='Times New Roman', color=colors.WHITE, size=24,bold=False,italic=False)

fill1 = PatternFill(start_color='00003366', fill_type='solid')
fill2 = PatternFill(start_color='000066CC', fill_type='solid')
for i, row in enumerate(cell_range):
    if i % 2 == 0:
        for cell in row:
            cell.fill = fill1
    else:
        for cell in row:
            cell.fill = fill2

cell_range = ws['A1:D1']
for row in cell_range:
    for cell in row:
        cell.alignment = xl.styles.Alignment(horizontal='center',vertical='bottom')
        cell.font = Font(name='Times New Roman', size=18,bold=True,italic=True)

coindesk = 'https://coindesk.com/data/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(coindesk, headers=headers)
page = urlopen(coindesk).read()			
soup = BeautifulSoup(page, 'html.parser')

prices = soup.findAll("h6", class_='typography__StyledTypography-owin6q-0 brrRIQ')
names = soup.findAll("span", class_='typography__StyledTypography-owin6q-0 gtxndB')
percentages = soup.findAll("div", class_='percentage')

AllImages = []
imgs = soup.findAll('img')
for i, image in enumerate(imgs):
    img_url = image['src']
    AllImages.append(img_url)

MyImages = AllImages[16:]
price = []
name = []
percent = []
for p in prices:
    price.append(p.text)
for n in names:
    name.append(n.text)
for p in percentages:
    percent.append(p.text)

price5 = price[:5]
name5 = name[:5]
percent5 = percent[:5]
image5 = MyImages[:5]

for i, url in enumerate(image5):
    temp_file, _ = urlretrieve(url)
    img = Image(temp_file)
    ws.add_image(img,f'A{i+2}')
for i, value in enumerate(name5):
    cell = ws.cell(row=i+2,column=2)
    cell.value = value
for i, value in enumerate(price5):
    cell = ws.cell(row=i+2,column=3)
    cell.value = value
for i, value in enumerate(percent5):
    cell = ws.cell(row=i+2,column=4)
    cell.value = value

wb.save('webscraping-project.xlsx')

_price = [float(x.replace(',','').strip('$')) for x in price5]
_percent = [float(x.strip('%'))*.01 for x in percent5]

dollar_change = [_price[i] * _percent[i] for i in range(len(_price))]
print(dollar_change)
y=0
for x in dollar_change:
    if x > 5:
        message = f'{name5[y]} price has gone up by more then $5!'

        client = Client(keys.accountSID,keys.auth_token)
        TWnumber = "+13204088142"
        myphone = "+15129130434"
        textmsg = client.messages.create(to=myphone,from_=TWnumber,body=message)
        print(textmsg.status)
    y=+1
    if x <-5:
        message = f'{name5[y]} price has gone down by more then $5.'

        client = Client(keys.accountSID,keys.auth_token)
        TWnumber = "+13204088142"
        myphone = "+15129130434"
        textmsg = client.messages.create(to=myphone,from_=TWnumber,body=message)
        print(textmsg.status)